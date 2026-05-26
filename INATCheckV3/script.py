import os
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook

# ---------------- UI ----------------

root = tk.Tk()
root.withdraw()

print("Select the Excel (.xlsx) file...")
xlsx_file = filedialog.askopenfilename(
    title="Select Excel File",
    filetypes=[("Excel Files", "*.xlsx")]
)

if not xlsx_file:
    print("No Excel file selected.")
    exit()

print("Select the folder containing the folders to check...")
main_folder = filedialog.askdirectory(
    title="Select Main Folder"
)

if not main_folder:
    print("No folder selected.")
    exit()

# ------------------------------------

print("\nLoading Excel file...")

# Load workbook
workbook = load_workbook(xlsx_file, data_only=True)
sheet = workbook.active

print("Reading Excel values...")

# Read all Excel values
excel_values = []

for row in sheet.iter_rows(values_only=True):
    for cell in row:
        if cell is not None:
            excel_values.append(str(cell).strip().upper())

print(f"Loaded {len(excel_values)} Excel values.\n")

# Process folders
for folder_name in os.listdir(main_folder):

    folder_path = os.path.join(main_folder, folder_name)

    # Only process folders
    if os.path.isdir(folder_path):

        # Folder name already contains the code
        cleaned = folder_name.lstrip("0")

        # Prevent empty string
        if cleaned == "":
            cleaned = "0"

        found = False

        for value in excel_values:
            if cleaned.upper() in value:
                found = True
                break

        if found:
            print(f"[FOUND]     {folder_name} -> {cleaned}")
        else:
            print(f"[NOT FOUND] {folder_name} -> {cleaned}")

print("\nFinished.")
input("Press Enter to close...")